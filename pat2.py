'''
    - File downloader
        You need to import for different file extension. Comment imports as per required.
        Code is for .xlsx and used openpyxl for hyperlink value and target
'''
import urllib
import pickle
import os
import pandas as pd
import openpyxl
from os import path
from pathlib import Path
import requests
import docx, pptx,fitz
# from pandas.core.indexes.numeric import Float64Index
import wget

def is_downloadable(url):
    """
    Does the url contain a downloadable resource
    """
    try:
        h = requests.head(url, allow_redirects=True)
        header = h.headers
        content_type = header.get('content-type')
        if 'text' in content_type.lower():
            return False
        if 'html' in content_type.lower():
            return False
        return True
    except Exception as e:
        print(e)



def file_extension(file):
    try:
        actual_file = os.path.splitext(file)
        print(actual_file, file, actual_file[1])
        fp = []
        if actual_file[1] == '.doc':
            fp = open(file, 'r' ,encoding = 'cp1252', errors = 'ignore')
        if actual_file[1] == '.docx':
            fp = docx.Document(file)
            for i in fp.paragraphs:
                print(i.text)
        if actual_file[1] == '.pdf':
            fp = fitz.open(file)
        if actual_file[1] == '.pptx':
            fp = pptx.Presentation(file)
        if actual_file[1] == '.xlsx' or 'xlsx' in actual_file[1]:
            # file = 'D:\\sorting_searching\\sd.xlsx'
            # fp = pd.read_excel('sd.xlsx',sheet_name='Sheet3', header=None)
            wb = openpyxl.load_workbook(file)
            fp = wb.get_sheet_by_name('Sheet3')            
        if actual_file[1] == '.txt':
            fp = open(file,'r')
            # with open('patona2.txt','r') as fd:
                # print(fd.read())
                # print(fd.readlines())
        return fp, actual_file[1]
    except Exception as e:
        print(e)


def process_file(file):
    try:
        # sdt = pd.read_excel(file,sheet_name='Sheet3')
        # print(len(sdt))
        # print("file",file)
        sd, file_ext = file_extension(file)
        print(sd, file_ext)
        if file_ext == '.xlsx':
            vs = []
            vst =[]
            for i in range(1,1000000):
                # print(sd.cell(row=i, column=1).value)
                if not sd.cell(row=i, column=1).value or sd.cell(row=i, column=1).value == None:
                    break
                if sd.cell(row=i, column=1).hyperlink:
                    vals = sd.cell(row=i, column=1).value.split()[0]
                    valst = sd.cell(row=i, column=1).hyperlink
                    print("valst.display",valst.display)
                    print("valst.target",valst.target)
                    vs.append(vals)
                    vst.append(valst.target)
                    print("vs vst",vs, vst)
                    
            return vs, vst
        elif file_ext == '.txt':
            fd = sd
            dicty = {}
            count= 0
            lt = []
            lin = fd.readlines()
            for line in lin:
                count += 1
                lt.append(line.split())
            sorted(lt, key =lambda x:x[1])
            # print(lt)
            return lt,file_ext
    except Exception as e:
        print(e)

# file = 'patona.txt'


def url_process(lt,extr):
    try:
        for down in range(len(extr)):
            url = extr[down]
            r = requests.get(url, allow_redirects=True)
            # print("req",r)
            if is_downloadable(url):
                open(lt[down]+'.txt', 'wb').write(r.content)
                print("saved with file name", lt[down]+'.txt')
    except Exception as e:
        print(e)

file = 'sd.xlsx'
# print("Enter file path:")
# file = input()
# print(file)
lt, extr = process_file(file)
url_process(lt, extr)

