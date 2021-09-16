'''
    - File downloader
        You need to import for different file extension. Comment imports as per required.
        Code is for .xlsx and used openpyxl for hyperlink value and target
'''
import urllib
import pickle
import os
import pandas as pd
import openpyxl
import requests
import docx, pptx,fitz
# from pandas.core.indexes.numeric import Float64Index
import wget

def is_downloadable(url):
    """
    Does the url contain a downloadable resource
    """
    try:
        h = requests.head(url, allow_redirects=True)
        header = h.headers
        content_type = header.get('content-type')
        if 'text' in content_type.lower():
            return False
        if 'html' in content_type.lower():
            return False
        return True
    except Exception as e:
        print(e)



def file_extension(file):
    try:
        actual_file = os.path.splitext(file)
        # print(actual_file)
        if actual_file[1] == '.doc':
            fp = open(file, 'r' ,encoding = 'cp1252', errors = 'ignore')
        if actual_file[1] == '.docx':
            fp = docx.Document(file)
            for i in fp.paragraphs:
                print(i.text)
        if actual_file[1] == '.pdf':
            fp = fitz.open(file)
        if actual_file[1] == '.pptx':
            fp = pptx.Presentation(file)
        if actual_file[1] == '.xlsx':
            # fp = pd.read_excel('sd.xlsx',sheet_name='Sheet3', header=None)
            wb = openpyxl.load_workbook('sd.xlsx')
            fp = wb['Sheet3']
            # fp = wb.get_sheet_by_name('Sheet3')
            
        if actual_file[1] == '.txt':
            fp = open('patona2.txt','r')
            # with open('patona2.txt','r') as fd:
                # print(fd.read())
                # print(fd.readlines())
        return fp, actual_file[1]
    except Exception as e:
        print(e)


def process_file(file):
    try:
        # sdt = pd.read_excel(file,sheet_name='Sheet3')
        # print(len(sdt))
        sd, file_ext = file_extension(file)
        print(sd, file_ext)
        if file_ext == '.xlsx':
            vs = []
            vst =[]
            for i in range(1,1000000):
                print(sd.cell(row=i, column=1).value)
                if not sd.cell(row=i, column=1).value or sd.cell(row=i, column=1).value == None:
                    break
                if sd.cell(row=i, column=1).hyperlink:
                    vals = sd.cell(row=i, column=1).value.split()[0]
                    valst = sd.cell(row=i, column=1).hyperlink
                    print("valst.display",valst.display)
                    print("valst.target",valst.target)
                    # print("valst.tooltip",valst._value)
                    # print(vals)
                    # print(valst.target)
                    vs.append(vals)
                    vst.append(valst.target)
                    print("vs vst",vs, vst)
                    # print("fd",sd.cell(row=i, column=1).hyperlink,type(sd.cell(row=i, column=1).hyperlink))
                    # print("fd",sd.cell(row=i, column=1).hyperlink.value)
            # sfg = sd['link2 2'].str.split(" ", n = 1, expand = True)
            # sfg = sd[0].str.split(" ", n = 1, expand = True)
            # sfg = sfg.sort_values(by=[1], ascending=True)
            # print(sfg)
            # print("s1",sfg[0].to_list())
            # print("s12",sfg[1].to_list())
            # return sfg[0].to_list(),file_ext
            return vs, vst
        elif file_ext == '.txt':
            fd = sd
            dicty = {}
            count= 0
            lt = []
            lin = fd.readlines()
            for line in lin:
                count += 1
                lt.append(line.split())
            sorted(lt, key =lambda x:x[1])
            # print(lt)
            return lt,file_ext
            # [['https://www.facebook.com/favicon.ico', '1'], ['link2', '2'], ['link2', '3'], ['link0', '5'], ['link4', '5'], ['link1', '6'], ['link3', '7'], ['link4', '8'], ['link1', '6'], ['link3', '7'], ['link3', '3'], ['link6', '5']]
    except Exception as e:
        print(e)

# file = 'patona.txt'
file = 'sd.xlsx'
lt, extr = process_file(file)
print(lt,extr)

def url_process(lt,extr):
    try:
        for down in range(len(extr)):
            url = extr[down]
            print(url)
            # print("sds",lt[down]+'.txt')
            r = requests.get(url, allow_redirects=True)
            print("req",r)
            if is_downloadable(url):
                # print("yes")
                # saving the file
                open(lt[down]+'.txt', 'wb').write(r.content)
                print("saved with file name", lt[down]+'.txt')
            # if '.txt' in url:
            #     print("enter txt")
            #     r = requests.get(url)
            #     filename = lt[down]+'.txt'
            #     with open(filename, 'wb') as f:
            #         f.write(r.content)
                # import urllib3
                # connection_pool = urllib3.PoolManager()
                # data = connection_pool.request('GET',url )
                # # response = urllib3.
                # #   .urlopen(url)
                # # data = response.read()
                # filename = lt[down]+'.txt'
                # file_ = open(filename, 'w')
                # file_.write(data)
                # file_.close()
                print("saved txt")
                # lt[down]+'.txt'
    except Exception as e:
        print(e)

url_process(lt, extr)
# print(r)
# print(r.content)
# print(is_downloadable(url))
from multiprocessing import Pool, cpu_count

print(cpu_count())

